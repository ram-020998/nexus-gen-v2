"""
Merge Report Excel Service

This service generates Excel reports for merge sessions with complexity
analysis, time estimates, and detailed change information. It orchestrates
data enrichment, complexity calculation, and Excel formatting to produce
comprehensive pre-merge assessment reports.

The service follows the existing OOP architecture and reuses patterns from
ExcelService for consistent styling and formatting.
"""
import json
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config
from config.report_config import ReportConfig
from services.merge_assistant.complexity_calculator_service import (
    ComplexityCalculatorService
)


logger = logging.getLogger(__name__)


class ReportExportService:
    """
    Service for generating Excel reports from merge sessions.

    This service:
    1. Retrieves all changes from a merge session
    2. Enriches changes with complexity and time estimates
    3. Generates human-readable change descriptions
    4. Extracts SAIL code changes with truncation
    5. Formats data into a styled Excel workbook
    6. Saves the report to the outputs directory

    The service reuses styling patterns from ExcelService and follows
    the existing OOP architecture.
    """

    def __init__(
        self,
        complexity_calculator: Optional[ComplexityCalculatorService] = None,
        config: Optional[ReportConfig] = None
    ):
        """
        Initialize the report export service.

        Args:
            complexity_calculator: Optional ComplexityCalculatorService
                                 instance. If None, creates a new one.
            config: Optional ReportConfig instance. If None, uses default.
        """
        self.complexity_calculator = (
            complexity_calculator or ComplexityCalculatorService()
        )
        self.config = config or ReportConfig
        logger.info("ReportExportService initialized")

    def generate_report(
        self,
        session_id: int,
        merge_service
    ) -> str:
        """
        Generate Excel report for a merge session.

        This is the main orchestration method that:
        1. Retrieves all changes from the session
        2. Builds enriched report data with complexity and descriptions
        3. Creates and formats the Excel workbook
        4. Saves the file to the outputs directory

        Args:
            session_id: Merge session ID
            merge_service: ThreeWayMergeService instance for data retrieval

        Returns:
            str: Path to the generated Excel file

        Raises:
            ValueError: If session not found or has no changes
            Exception: If Excel generation fails

        Example:
            >>> service = ReportExportService()
            >>> path = service.generate_report(1, merge_service)
            >>> print(f"Report saved to: {path}")
        """
        logger.info(f"Starting report generation for session {session_id}")

        # Get session information
        session = merge_service.get_session(session_id)
        if not session:
            raise ValueError(f"Session {session_id} not found")

        # Get all changes using simple query (without heavy joinedloads)
        # This is much faster for report generation
        from models import Change
        changes_query = Change.query.filter(
            Change.session_id == session_id,
            Change.classification.in_(['NO_CONFLICT', 'CONFLICT', 'REMOVED_BUT_CUSTOMIZED'])
        ).order_by(Change.display_order).all()

        # Convert to dictionaries
        all_changes = [merge_service._build_change_dict(change) for change in changes_query]

        if not all_changes:
            raise ValueError(
                f"No changes found for session {session_id}"
            )

        logger.info(
            f"Retrieved {len(all_changes)} changes for report generation"
        )

        # Build enriched report data
        report_data = self._build_report_data(all_changes, merge_service)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Merge Assessment Report"

        # Format Excel file
        self._format_excel_file(ws, report_data, session)

        # Generate filename and save
        filename = self._generate_filename(session)
        output_path = Config.get_output_path(filename)

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(str(output_path))

        logger.info(f"Report generated successfully: {output_path}")
        return str(output_path)

    def _build_report_data(
        self,
        changes: List[Dict[str, Any]],
        merge_service
    ) -> List[Dict[str, Any]]:
        """
        Build enriched data for the report.

        Enriches each change with:
        - Complexity level
        - Estimated time
        - Human-readable change description
        - SAIL code changes (truncated if needed)
        - Serial number for display

        Args:
            changes: List of change dictionaries from merge service
            merge_service: ThreeWayMergeService for additional data

        Returns:
            List of enriched change dictionaries ready for Excel export
        """
        logger.info(f"Building report data for {len(changes)} changes")
        report_data = []

        for idx, change in enumerate(changes, start=1):
            # Extract objects for complexity calculation
            base_obj = change.get('base_object')
            customer_obj = change.get('customer_object')
            vendor_obj = change.get('vendor_object')

            # Calculate complexity
            complexity = self.complexity_calculator.calculate_complexity(
                change,
                base_obj,
                customer_obj,
                vendor_obj
            )

            # Calculate estimated time
            time_minutes = (
                self.complexity_calculator.calculate_estimated_time(
                    complexity
                )
            )
            time_display = (
                self.complexity_calculator.format_time_display(
                    time_minutes
                )
            )

            # Generate change description
            description = self._generate_change_description(change)

            # Extract SAIL changes
            sail_changes = self._extract_sail_changes(
                base_obj,
                customer_obj,
                vendor_obj
            )

            # Build enriched change data
            enriched_change = {
                'serial_number': idx,
                'category': self._format_classification(
                    change.get('classification', '')
                ),
                'object_name': change.get('name', ''),
                'object_uuid': change.get('uuid', ''),
                'change_description': description,
                'sail_changes': sail_changes,
                'complexity': complexity,
                'estimated_time': time_display,
                'comments': ''  # Reserved for future use
            }

            report_data.append(enriched_change)

        logger.info("Report data built successfully")
        return report_data

    def _format_excel_file(
        self,
        ws,
        report_data: List[Dict[str, Any]],
        session
    ) -> None:
        """
        Apply formatting to Excel file.

        Creates headers, populates data rows, and applies styling
        following the patterns from ExcelService (purple theme).

        Args:
            ws: openpyxl worksheet
            report_data: List of enriched change dictionaries
            session: MergeSession object for metadata
        """
        # Define styles (reusing ExcelService patterns)
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="8B5CF6",
            end_color="8B5CF6",
            fill_type="solid"
        )
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title row
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = (
            f"Merge Assessment Report - {session.reference_id}"
        )
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Add metadata row
        ws.merge_cells('A2:I2')
        metadata_cell = ws['A2']
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        metadata_cell.value = (
            f"Generated: {timestamp} | "
            f"Base: {session.base_package_name} | "
            f"Customized: {session.customized_package_name} | "
            f"New Vendor: {session.new_vendor_package_name}"
        )
        metadata_cell.font = Font(size=10, italic=True)
        metadata_cell.alignment = Alignment(horizontal='center')

        # Add headers in row 4
        headers = self.config.EXCEL_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

        # Add data rows starting from row 5
        for row_idx, change in enumerate(report_data, start=5):
            ws.cell(
                row=row_idx,
                column=1,
                value=change['serial_number']
            ).border = border

            ws.cell(
                row=row_idx,
                column=2,
                value=change['category']
            ).border = border

            ws.cell(
                row=row_idx,
                column=3,
                value=change['object_name']
            ).border = border

            ws.cell(
                row=row_idx,
                column=4,
                value=change['object_uuid']
            ).border = border

            # Change description with wrap text
            desc_cell = ws.cell(
                row=row_idx,
                column=5,
                value=change['change_description']
            )
            desc_cell.border = border
            desc_cell.alignment = Alignment(
                wrap_text=True,
                vertical='top'
            )

            # SAIL changes with wrap text
            sail_cell = ws.cell(
                row=row_idx,
                column=6,
                value=change['sail_changes']
            )
            sail_cell.border = border
            sail_cell.alignment = Alignment(
                wrap_text=True,
                vertical='top'
            )

            ws.cell(
                row=row_idx,
                column=7,
                value=change['complexity']
            ).border = border

            ws.cell(
                row=row_idx,
                column=8,
                value=change['estimated_time']
            ).border = border

            ws.cell(
                row=row_idx,
                column=9,
                value=change['comments']
            ).border = border

        # Adjust column widths
        column_widths = [10, 20, 30, 40, 40, 50, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Set row heights for better readability
        ws.row_dimensions[1].height = 25  # Title
        ws.row_dimensions[2].height = 20  # Metadata
        ws.row_dimensions[4].height = 30  # Headers
        for row_num in range(5, 5 + len(report_data)):
            ws.row_dimensions[row_num].height = 60

    def _generate_change_description(
        self,
        change: Dict[str, Any]
    ) -> str:
        """
        Generate human-readable change description.

        Creates descriptions based on change type and classification:
        - Added: "New object added in vendor release"
        - Removed: "Object removed in vendor release"
        - Modified: Describes specific changes
        - Conflict: Describes conflicting changes from both sides

        Args:
            change: Change dictionary

        Returns:
            str: Human-readable description
        """
        change_type = change.get('change_type', '')
        classification = change.get('classification', '')
        vendor_change = change.get('vendor_change_type', '')
        customer_change = change.get('customer_change_type', '')

        # Handle added objects
        if change_type == 'ADDED' or vendor_change == 'ADDED':
            return "New object added in vendor release"

        # Handle removed objects
        if change_type == 'REMOVED' or vendor_change == 'REMOVED':
            if classification == 'REMOVED_BUT_CUSTOMIZED':
                return (
                    "Object removed in vendor release but was "
                    "customized in customer version"
                )
            return "Object removed in vendor release"

        # Handle conflicts
        if classification == 'CONFLICT':
            vendor_desc = self._describe_change_type(vendor_change)
            customer_desc = self._describe_change_type(customer_change)
            return (
                f"Conflicting changes: Vendor {vendor_desc}, "
                f"Customer {customer_desc}"
            )

        # Handle modifications
        if change_type == 'MODIFIED' or vendor_change == 'MODIFIED':
            # Try to get specific field changes
            base_obj = change.get('base_object')
            vendor_obj = change.get('vendor_object')

            if base_obj and vendor_obj:
                field_changes = self._extract_field_changes(
                    base_obj,
                    vendor_obj
                )
                if field_changes and field_changes != "No changes detected":
                    return f"Modified: {field_changes}"

            return "Object modified in vendor release"

        # Default description
        return "Object changed"

    def _describe_change_type(self, change_type: str) -> str:
        """
        Convert change type to human-readable description.

        Args:
            change_type: Change type code

        Returns:
            str: Human-readable description
        """
        descriptions = {
            'ADDED': 'added',
            'REMOVED': 'removed',
            'MODIFIED': 'modified',
            'SAIL_CODE_CHANGED': 'SAIL code changed',
            'BUSINESS_LOGIC_CHANGED': 'business logic changed',
            'FIELDS_CHANGED': 'fields changed'
        }
        return descriptions.get(change_type, 'changed')

    def _extract_sail_changes(
        self,
        base_object: Optional[Any],
        customer_object: Optional[Any],
        vendor_object: Optional[Any]
    ) -> str:
        """
        Extract and format SAIL code changes.

        For Interface and Expression Rule objects, extracts SAIL code
        differences. For other objects, falls back to field changes.
        Truncates output if it exceeds the configured maximum length.

        Args:
            base_object: Base version object
            customer_object: Customer version object
            vendor_object: Vendor version object

        Returns:
            str: Formatted SAIL changes or field changes
        """
        # Extract SAIL code
        base_sail = self._get_sail_code(base_object)
        vendor_sail = self._get_sail_code(vendor_object)

        # If both have SAIL code, show the difference
        if base_sail and vendor_sail:
            # Simple diff: show new version with indication
            diff_text = f"SAIL Code Updated:\n{vendor_sail}"
            return self._truncate_text(diff_text)
        elif vendor_sail and not base_sail:
            # New SAIL code added
            diff_text = f"SAIL Code Added:\n{vendor_sail}"
            return self._truncate_text(diff_text)
        elif base_sail and not vendor_sail:
            # SAIL code removed
            return "SAIL code removed"

        # No SAIL code changes, try field changes
        field_changes = self._extract_field_changes(
            base_object,
            vendor_object
        )
        return field_changes

    def _extract_field_changes(
        self,
        base_object: Optional[Any],
        new_object: Optional[Any]
    ) -> str:
        """
        Extract and format field changes for non-SAIL objects.

        Compares fields between base and new versions and generates
        a summary of what changed.

        Args:
            base_object: Base version object
            new_object: New version object

        Returns:
            str: Summary of field changes
        """
        if not base_object or not new_object:
            return "No changes detected"

        # Extract fields
        base_fields = self._get_fields(base_object)
        new_fields = self._get_fields(new_object)

        if not base_fields and not new_fields:
            return "No field data available"

        # Find changed fields
        changed_fields = []

        # Check for added/modified fields
        for field_name, new_value in new_fields.items():
            if field_name not in base_fields:
                changed_fields.append(f"Added: {field_name}")
            elif base_fields[field_name] != new_value:
                changed_fields.append(f"Modified: {field_name}")

        # Check for removed fields
        for field_name in base_fields:
            if field_name not in new_fields:
                changed_fields.append(f"Removed: {field_name}")

        if not changed_fields:
            return "No changes detected"

        # Format as comma-separated list
        changes_text = ", ".join(changed_fields[:5])  # Limit to 5 fields
        if len(changed_fields) > 5:
            changes_text += f" (and {len(changed_fields) - 5} more)"

        return changes_text

    def _get_sail_code(self, obj: Optional[Any]) -> Optional[str]:
        """
        Extract SAIL code from an object.

        Handles both dictionary and SQLAlchemy model objects.

        Args:
            obj: Object containing SAIL code

        Returns:
            str or None: SAIL code if present
        """
        if obj is None:
            return None

        # Handle dictionary
        if isinstance(obj, dict):
            return obj.get('sail_code')

        # Handle model
        if hasattr(obj, 'sail_code'):
            return obj.sail_code

        return None

    def _get_fields(self, obj: Optional[Any]) -> Dict[str, Any]:
        """
        Extract fields from an object.

        Handles both dictionary and SQLAlchemy model objects.

        Args:
            obj: Object containing fields

        Returns:
            dict: Fields dictionary
        """
        if obj is None:
            return {}

        # Handle dictionary
        if isinstance(obj, dict):
            fields = obj.get('fields')
            if isinstance(fields, str):
                try:
                    return json.loads(fields)
                except json.JSONDecodeError:
                    return {}
            return fields or {}

        # Handle model
        if hasattr(obj, 'fields'):
            fields = obj.fields
            if isinstance(fields, str):
                try:
                    return json.loads(fields)
                except json.JSONDecodeError:
                    return {}
            return fields or {}

        return {}

    def _truncate_text(self, text: str) -> str:
        """
        Truncate text to maximum length with ellipsis.

        Args:
            text: Text to truncate

        Returns:
            str: Truncated text with ellipsis if needed
        """
        max_length = self.config.SAIL_CODE_MAX_LENGTH
        if len(text) <= max_length:
            return text

        return (
            text[:max_length] +
            self.config.SAIL_CODE_TRUNCATION_SUFFIX
        )

    def _format_classification(self, classification: str) -> str:
        """
        Format classification for display.

        Converts internal classification codes to human-readable labels.

        Args:
            classification: Internal classification code

        Returns:
            str: Human-readable classification label
        """
        labels = {
            'NO_CONFLICT': 'No Conflict',
            'CONFLICT': 'Conflicting',
            'CUSTOMER_ONLY': 'Customer Only',
            'REMOVED_BUT_CUSTOMIZED': 'Deprecated'
        }
        return labels.get(classification, classification)

    def _generate_filename(self, session) -> str:
        """
        Generate filename for the Excel report.

        Args:
            session: MergeSession object

        Returns:
            str: Filename for the report
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return (
            f"merge_report_{session.reference_id}_{timestamp}.xlsx"
        )
