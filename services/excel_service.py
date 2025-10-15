"""
Excel Export Service - Generate Excel files from breakdown data
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import json
from config import Config

class ExcelService:
    """Handle Excel file generation"""
    
    def __init__(self):
        self.config = Config
    
    def create_breakdown_excel(self, request_id: int, breakdown_data: dict, filename: str = None) -> str:
        """Create Excel file from breakdown data"""
        try:
            # Create output directory
            output_dir = self.config.OUTPUT_FOLDER / str(request_id)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename
            if not filename:
                filename = f"breakdown_{request_id}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename = f"{filename.rsplit('.', 1)[0]}.xlsx"
            
            excel_path = output_dir / filename
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Spec Breakdown"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Headers
            headers = ["Epic", "Story Name", "Acceptance Criteria", "Issue Type", "Points"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            row = 2
            epic = breakdown_data.get('epic', 'N/A')
            
            for story in breakdown_data.get('stories', []):
                ws.cell(row=row, column=1, value=epic).border = border
                ws.cell(row=row, column=2, value=story.get('story_name', '')).border = border
                
                # Format acceptance criteria
                ac_cell = ws.cell(row=row, column=3, value=story.get('acceptance_criteria', ''))
                ac_cell.border = border
                ac_cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                ws.cell(row=row, column=4, value=story.get('issue_type', '')).border = border
                ws.cell(row=row, column=5, value=story.get('points', '')).border = border
                
                row += 1
            
            # Adjust column widths
            column_widths = [15, 30, 50, 15, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Set row heights for better readability
            for row_num in range(2, row):
                ws.row_dimensions[row_num].height = 60
            
            # Save workbook
            wb.save(str(excel_path))
            
            return str(excel_path)
            
        except Exception as e:
            print(f"Excel generation error: {e}")
            raise e
    
    def create_design_excel(self, request_id: int, design_data: dict, filename: str = None) -> str:
        """Create Excel file from design document data"""
        try:
            output_dir = self.config.OUTPUT_FOLDER / str(request_id)
            output_dir.mkdir(parents=True, exist_ok=True)
            
            if not filename:
                filename = f"design_{request_id}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename = f"{filename.rsplit('.', 1)[0]}.xlsx"
            
            excel_path = output_dir / filename
            
            wb = openpyxl.Workbook()
            
            # Overview sheet
            ws_overview = wb.active
            ws_overview.title = "Overview"
            
            design_doc = design_data.get('design_document', {})
            
            ws_overview.cell(row=1, column=1, value="Design Document Overview").font = Font(bold=True, size=14)
            ws_overview.cell(row=3, column=1, value="Description:")
            ws_overview.cell(row=3, column=2, value=design_doc.get('overview', ''))
            
            # Objects sheet
            ws_objects = wb.create_sheet("Objects & Components")
            
            headers = ["Name", "Type", "Description", "Methods"]
            for col, header in enumerate(headers, 1):
                cell = ws_objects.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            row = 2
            for obj in design_doc.get('objects', []):
                ws_objects.cell(row=row, column=1, value=obj.get('name', ''))
                ws_objects.cell(row=row, column=2, value=obj.get('type', ''))
                ws_objects.cell(row=row, column=3, value=obj.get('description', ''))
                ws_objects.cell(row=row, column=4, value=', '.join(obj.get('methods', [])))
                row += 1
            
            # Implementation Notes sheet
            ws_notes = wb.create_sheet("Implementation Notes")
            ws_notes.cell(row=1, column=1, value="Implementation Notes").font = Font(bold=True)
            
            row = 2
            for note in design_doc.get('implementation_notes', []):
                ws_notes.cell(row=row, column=1, value=f"• {note}")
                row += 1
            
            # Dependencies sheet
            ws_deps = wb.create_sheet("Dependencies")
            ws_deps.cell(row=1, column=1, value="Dependencies").font = Font(bold=True)
            
            row = 2
            for dep in design_doc.get('dependencies', []):
                ws_deps.cell(row=row, column=1, value=f"• {dep}")
                row += 1
            
            wb.save(str(excel_path))
            return str(excel_path)
            
        except Exception as e:
            print(f"Excel generation error: {e}")
            raise e
