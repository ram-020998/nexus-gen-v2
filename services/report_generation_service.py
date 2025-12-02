"""
Report Generation Service

Generates comprehensive merge reports in PDF and Word formats.
Includes session information, statistics, and detailed change lists.
"""

import logging
import os
from datetime import datetime
from typing import Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

from core.base_service import BaseService
from models import db, MergeSession, Change, Package, ObjectVersion


class ReportGenerationService(BaseService):
    """
    Service for generating merge reports.

    Provides methods for:
    - Generating Excel reports
    - Caching generated reports
    - Including comprehensive session data

    Reports include:
    - Session Information
    - Package Information
    - Statistics (by classification, by object type)
    - Detailed Change List with SAIL code

    Example:
        >>> service = ReportGenerationService()
        >>> report_path = service.generate_report('MS_ABC123', format='xlsx')
        >>> print(f"Report generated: {report_path}")
    """

    def __init__(self, container=None):
        """Initialize service with dependencies."""
        super().__init__(container)
        self.logger = logging.getLogger(__name__)
        self.reports_dir = 'outputs/reports'
        self._ensure_reports_directory()

    def _initialize_dependencies(self) -> None:
        """Initialize service dependencies."""
        pass

    def _ensure_reports_directory(self) -> None:
        """Ensure the reports directory exists."""
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
            self.logger.info(f"Created reports directory: {self.reports_dir}")

    def generate_report(
        self,
        reference_id: str,
        format: str = 'xlsx'
    ) -> str:
        """
        Generate a merge report.

        Args:
            reference_id: Session reference ID
            format: Report format ('xlsx' only)

        Returns:
            Path to generated report file

        Raises:
            ValueError: If session not found or format invalid

        Example:
            >>> report_path = service.generate_report(
            ...     'MS_ABC123', format='xlsx'
            ... )
        """
        # Validate format
        if format not in ['xlsx']:
            raise ValueError(
                f"Invalid format: {format}. Must be 'xlsx'"
            )

        # Check cache first
        cached_path = self._get_cached_report_path(reference_id, format)
        if cached_path and os.path.exists(cached_path):
            # Check if cache is still valid (less than 1 hour old)
            cache_age = (
                datetime.now().timestamp() - os.path.getmtime(cached_path)
            )
            if cache_age < 3600:  # 1 hour in seconds
                self.logger.info(
                    f"Using cached report for {reference_id} "
                    f"(age: {int(cache_age/60)} minutes)"
                )
                return cached_path

        # Get session
        session = db.session.query(MergeSession).filter_by(
            reference_id=reference_id
        ).first()

        if not session:
            raise ValueError(f"Session not found: {reference_id}")

        # Gather report data
        data = self._gather_report_data(session)

        # Generate Excel report
        report_path = self._generate_excel_report(session, data)

        self.logger.info(
            f"Generated {format.upper()} report for {reference_id}: "
            f"{report_path}"
        )

        return report_path

    def _get_cached_report_path(
        self,
        reference_id: str,
        format: str
    ) -> str:
        """
        Get the path to a cached report.

        Args:
            reference_id: Session reference ID
            format: Report format

        Returns:
            Path to cached report file
        """
        filename = f"{reference_id}_report.{format}"
        return os.path.join(self.reports_dir, filename)

    def _gather_report_data(self, session: MergeSession) -> Dict[str, Any]:
        """
        Gather all data needed for report generation.

        Args:
            session: Merge session

        Returns:
            Dict containing all report data
        """
        # Get packages
        packages = db.session.query(Package).filter_by(
            session_id=session.id
        ).all()

        package_info = {}
        for pkg in packages:
            package_info[pkg.package_type] = {
                'filename': pkg.filename,
                'total_objects': pkg.total_objects
            }

        # Get changes with object details
        changes = db.session.query(Change).filter_by(
            session_id=session.id
        ).order_by(Change.display_order).all()

        # Get package references
        base_pkg = next(
            (p for p in packages if p.package_type == 'base'), None
        )
        customized_pkg = next(
            (p for p in packages if p.package_type == 'customized'), None
        )
        vendor_pkg = next(
            (p for p in packages if p.package_type == 'new_vendor'), None
        )

        change_list = []
        for change in changes:
            obj = change.object

            # Get versions from all packages
            base_version = None
            customized_version = None
            vendor_version = None

            if base_pkg:
                base_version = db.session.query(ObjectVersion).filter_by(
                    object_id=obj.id, package_id=base_pkg.id
                ).first()

            if customized_pkg:
                customized_version = db.session.query(
                    ObjectVersion
                ).filter_by(
                    object_id=obj.id, package_id=customized_pkg.id
                ).first()

            if vendor_pkg:
                vendor_version = db.session.query(ObjectVersion).filter_by(
                    object_id=obj.id, package_id=vendor_pkg.id
                ).first()

            # Generate meaningful change description
            actual_change = self._generate_actual_change_description(
                change, obj, base_version, customized_version, vendor_version
            )

            change_list.append({
                'display_order': change.display_order,
                'object_name': obj.name,
                'object_uuid': obj.uuid,
                'object_type': obj.object_type,
                'description': obj.description,
                'classification': change.classification,
                'vendor_change_type': change.vendor_change_type,
                'customer_change_type': change.customer_change_type,
                'status': change.status,
                'notes': change.notes,
                'actual_change': actual_change,
                'ai_summary': change.ai_summary,
                'ai_summary_status': change.ai_summary_status
            })

        # Calculate statistics
        stats = self._calculate_statistics(session, changes)

        return {
            'session': session,
            'packages': package_info,
            'changes': change_list,
            'statistics': stats
        }

    def _calculate_statistics(
        self,
        session: MergeSession,
        changes: list
    ) -> Dict[str, Any]:
        """
        Calculate statistics for the report.

        Args:
            session: Merge session
            changes: List of Change objects

        Returns:
            Dict containing statistics
        """
        # Classification breakdown
        by_classification = {}
        classifications = ['NO_CONFLICT', 'CONFLICT', 'NEW', 'DELETED']
        for classification in classifications:
            count = sum(
                1 for c in changes if c.classification == classification
            )
            by_classification[classification] = count

        # Object type breakdown
        by_object_type = {}
        for change in changes:
            obj_type = change.object.object_type
            by_object_type[obj_type] = by_object_type.get(obj_type, 0) + 1

        # Status breakdown
        by_status = {}
        for status in ['pending', 'reviewed', 'skipped']:
            count = sum(1 for c in changes if c.status == status)
            by_status[status] = count

        return {
            'total_changes': session.total_changes,
            'reviewed_count': session.reviewed_count,
            'skipped_count': session.skipped_count,
            'pending_count': (
                session.total_changes -
                session.reviewed_count -
                session.skipped_count
            ),
            'by_classification': by_classification,
            'by_object_type': by_object_type,
            'by_status': by_status,
            'estimated_complexity': session.estimated_complexity,
            'estimated_time_hours': session.estimated_time_hours
        }

    def _generate_excel_report(
        self,
        session: MergeSession,
        data: Dict[str, Any]
    ) -> str:
        """
        Generate Excel report using openpyxl.

        Args:
            session: Merge session
            data: Report data

        Returns:
            Path to generated Excel file
        """
        # Create workbook
        wb = Workbook()

        # Create sheets
        self._create_summary_sheet(wb, session, data)
        self._create_changes_sheet(wb, session, data)

        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Save workbook
        report_path = self._get_cached_report_path(
            session.reference_id,
            'xlsx'
        )
        wb.save(report_path)

        return report_path

    def _create_summary_sheet(
        self,
        wb: Workbook,
        session: MergeSession,
        data: Dict[str, Any]
    ) -> None:
        """Create summary sheet with modern formatting."""
        ws = wb.active
        ws.title = 'Summary'

        # Define modern color scheme
        primary_color = '366092'  # Dark blue
        secondary_color = '5B9BD5'  # Light blue
        accent_color = 'F4B183'  # Orange
        success_color = '70AD47'  # Green
        warning_color = 'FFC000'  # Yellow
        danger_color = 'C00000'  # Red
        light_gray = 'F2F2F2'  # Light gray background

        # Define border styles
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Title Section with modern styling
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = '📊 Three-Way Merge Analysis Report'
        title_cell.font = Font(
            bold=True, size=20, color='FFFFFF', name='Calibri'
        )
        title_cell.fill = PatternFill(
            start_color=primary_color,
            end_color=primary_color,
            fill_type='solid'
        )
        title_cell.alignment = Alignment(
            horizontal='center', vertical='center'
        )
        ws.row_dimensions[1].height = 35

        # Subtitle with timestamp
        ws.merge_cells('A2:F2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = (
            f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        )
        subtitle_cell.font = Font(size=11, italic=True, color='666666')
        subtitle_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        row = 4

        # Session Information Section
        self._add_section_header(
            ws, row, 'A', 'F', '🔍 Session Information', primary_color
        )
        row += 1

        session_data = [
            ('Reference ID', session.reference_id, None),
            ('Status', session.status.upper(), self._get_status_color(
                session.status
            )),
            (
                'Created',
                session.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                None
            ),
            ('Total Changes', session.total_changes, secondary_color),
            (
                'Estimated Complexity',
                session.estimated_complexity or 'N/A',
                None
            ),
            (
                'Estimated Time',
                f"{session.estimated_time_hours or 0} hours",
                None
            )
        ]

        for label, value, bg_color in session_data:
            self._add_data_row(
                ws, row, label, value, thin_border, bg_color
            )
            row += 1

        row += 1

        # Package Information Section
        self._add_section_header(
            ws, row, 'A', 'F', '📦 Package Information', primary_color
        )
        row += 1

        packages = data['packages']
        pkg_types = [
            ('base', 'Base Version', secondary_color),
            ('customized', 'Customer Version', accent_color),
            ('new_vendor', 'New Vendor Version', success_color)
        ]

        for pkg_type, display_name, color in pkg_types:
            if pkg_type in packages:
                pkg = packages[pkg_type]
                # Package header
                ws.merge_cells(f'A{row}:F{row}')
                pkg_cell = ws[f'A{row}']
                pkg_cell.value = display_name
                pkg_cell.font = Font(bold=True, size=12, color='FFFFFF')
                pkg_cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type='solid'
                )
                pkg_cell.alignment = Alignment(
                    horizontal='left', vertical='center'
                )
                pkg_cell.border = thin_border
                ws.row_dimensions[row].height = 25
                row += 1

                # Package details
                self._add_data_row(
                    ws, row, '  Filename', pkg['filename'], thin_border
                )
                row += 1
                self._add_data_row(
                    ws, row, '  Total Objects',
                    pkg['total_objects'], thin_border
                )
                row += 1

        row += 1

        # Statistics Section
        self._add_section_header(
            ws, row, 'A', 'F', '📈 Statistics & Metrics', primary_color
        )
        row += 1

        stats = data['statistics']

        # Progress subsection
        ws.merge_cells(f'A{row}:F{row}')
        progress_cell = ws[f'A{row}']
        progress_cell.value = 'Review Progress'
        progress_cell.font = Font(bold=True, size=11)
        progress_cell.fill = PatternFill(
            start_color=light_gray, end_color=light_gray, fill_type='solid'
        )
        progress_cell.border = thin_border
        row += 1

        progress_data = [
            ('Total Changes', stats['total_changes'], None),
            ('Reviewed', stats['reviewed_count'], success_color),
            ('Skipped', stats['skipped_count'], warning_color),
            ('Pending', stats['pending_count'], danger_color)
        ]

        for label, value, bg_color in progress_data:
            self._add_data_row(
                ws, row, f'  {label}', value, thin_border, bg_color
            )
            row += 1

        row += 1

        # Classification breakdown
        ws.merge_cells(f'A{row}:F{row}')
        class_cell = ws[f'A{row}']
        class_cell.value = 'Classification Breakdown'
        class_cell.font = Font(bold=True, size=11)
        class_cell.fill = PatternFill(
            start_color=light_gray, end_color=light_gray, fill_type='solid'
        )
        class_cell.border = thin_border
        row += 1

        classifications = [
            ('NO_CONFLICT', 'No Conflict', success_color),
            ('CONFLICT', 'Conflict', danger_color),
            ('NEW', 'New', secondary_color),
            ('DELETED', 'Deleted', warning_color)
        ]

        for key, label, bg_color in classifications:
            count = stats['by_classification'].get(key, 0)
            self._add_data_row(
                ws, row, f'  {label}', count, thin_border, bg_color
            )
            row += 1

        row += 1

        # Object type breakdown
        ws.merge_cells(f'A{row}:F{row}')
        obj_type_cell = ws[f'A{row}']
        obj_type_cell.value = 'Object Type Distribution'
        obj_type_cell.font = Font(bold=True, size=11)
        obj_type_cell.fill = PatternFill(
            start_color=light_gray, end_color=light_gray, fill_type='solid'
        )
        obj_type_cell.border = thin_border
        row += 1

        obj_types = sorted(
            stats['by_object_type'].items(),
            key=lambda x: x[1],
            reverse=True
        )

        for obj_type, count in obj_types:
            self._add_data_row(
                ws, row, f'  {obj_type}', count, thin_border
            )
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    def _add_section_header(
        self, ws, row, start_col, end_col, title, color
    ):
        """Add a styled section header."""
        ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
        header_cell = ws[f'{start_col}{row}']
        header_cell.value = title
        header_cell.font = Font(
            bold=True, size=14, color='FFFFFF', name='Calibri'
        )
        header_cell.fill = PatternFill(
            start_color=color, end_color=color, fill_type='solid'
        )
        header_cell.alignment = Alignment(
            horizontal='left', vertical='center'
        )
        ws.row_dimensions[row].height = 30

    def _add_data_row(
        self, ws, row, label, value, border, bg_color=None
    ):
        """Add a data row with consistent styling."""
        # Label cell
        label_cell = ws[f'A{row}']
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10)
        label_cell.border = border
        label_cell.alignment = Alignment(vertical='center')

        # Value cell
        ws.merge_cells(f'B{row}:F{row}')
        value_cell = ws[f'B{row}']
        value_cell.value = value
        value_cell.font = Font(size=10)
        value_cell.border = border
        value_cell.alignment = Alignment(vertical='center')

        if bg_color:
            value_cell.fill = PatternFill(
                start_color=bg_color, end_color=bg_color, fill_type='solid'
            )
            value_cell.font = Font(size=10, bold=True, color='FFFFFF')

        ws.row_dimensions[row].height = 20

    def _get_status_color(self, status):
        """Get color based on status."""
        status_colors = {
            'ready': '70AD47',  # Green
            'in_progress': 'FFC000',  # Yellow
            'completed': '5B9BD5',  # Blue
            'error': 'C00000'  # Red
        }
        return status_colors.get(status.lower(), '5B9BD5')

    def _generate_actual_change_description(
        self,
        change,
        obj,
        base_version,
        customized_version,
        vendor_version
    ):
        """
        Generate a meaningful description of what actually changed.

        Args:
            change: Change object
            obj: Object from object_lookup
            base_version: Base package version
            customized_version: Customer package version
            vendor_version: New vendor package version

        Returns:
            String describing the actual change
        """
        vendor_change = change.vendor_change_type
        customer_change = change.customer_change_type
        classification = change.classification

        # For NEW objects
        if classification == 'NEW':
            if vendor_version and vendor_version.sail_code:
                code_preview = vendor_version.sail_code[:500]
                return f"NEW: Added in vendor version\n\n{code_preview}"
            return "NEW: Object added in vendor version"

        # For DELETED objects
        if classification == 'DELETED':
            if base_version and base_version.sail_code:
                code_preview = base_version.sail_code[:500]
                return (
                    f"DELETED: Removed in vendor version\n\n"
                    f"Original code:\n{code_preview}"
                )
            return "DELETED: Object removed in vendor version"

        # For MODIFIED objects - show diff
        if vendor_change == 'MODIFIED':
            # Try to generate SAIL diff
            if base_version and vendor_version:
                if base_version.sail_code and vendor_version.sail_code:
                    from services.sail_diff_service import (
                        SailDiffService
                    )
                    diff_service = SailDiffService(context_lines=2)
                    hunks = diff_service.generate_unified_diff(
                        base_version.sail_code,
                        vendor_version.sail_code,
                        "Base",
                        "New Vendor"
                    )

                    if hunks:
                        # Format diff as text
                        diff_text = self._format_diff_hunks(hunks)
                        return (
                            f"MODIFIED: SAIL code changed\n\n"
                            f"{diff_text[:800]}"
                        )

                # Check for field changes
                if base_version.fields and vendor_version.fields:
                    import json
                    is_str = isinstance(base_version.fields, str)
                    base_fields = (
                        json.loads(base_version.fields)
                        if is_str else base_version.fields
                    )
                    is_str = isinstance(vendor_version.fields, str)
                    vendor_fields = (
                        json.loads(vendor_version.fields)
                        if is_str else vendor_version.fields
                    )

                    if base_fields != vendor_fields:
                        base_str = str(base_fields)[:200]
                        vendor_str = str(vendor_fields)[:200]
                        return (
                            f"MODIFIED: Fields/properties changed\n"
                            f"Base: {base_str}\n"
                            f"Vendor: {vendor_str}"
                        )

            return "MODIFIED: Object modified in vendor version"

        # For DEPRECATED objects
        if vendor_change == 'DEPRECATED':
            return (
                "DEPRECATED: Object marked as deprecated "
                "in vendor version"
            )

        # Default
        return (
            f"Change Type: Vendor={vendor_change}, "
            f"Customer={customer_change}"
        )

    def _format_diff_hunks(self, hunks):
        """Format diff hunks as readable text."""
        lines = []
        for hunk in hunks[:3]:  # Limit to first 3 hunks
            lines.append(
                f"@@ -{hunk.old_start},{hunk.old_count} "
                f"+{hunk.new_start},{hunk.new_count} @@"
            )
            for line in hunk.lines[:20]:  # Limit lines per hunk
                if line.line_type.value == 'added':
                    lines.append(f"+ {line.content}")
                elif line.line_type.value == 'removed':
                    lines.append(f"- {line.content}")
                elif line.line_type.value == 'unchanged':
                    lines.append(f"  {line.content}")

        return '\n'.join(lines)

    def _create_changes_sheet(
        self,
        wb: Workbook,
        session: MergeSession,
        data: Dict[str, Any]
    ) -> None:
        """Create changes sheet with modern formatting."""
        ws = wb.create_sheet('Changes')

        # Define colors
        header_color = '366092'
        conflict_color = 'FFE6E6'  # Light red
        no_conflict_color = 'E6F4EA'  # Light green
        new_color = 'E3F2FD'  # Light blue
        deleted_color = 'FFF4E6'  # Light orange
        alt_row_color = 'F8F9FA'  # Very light gray

        # Define border
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Header style
        header_fill = PatternFill(
            start_color=header_color,
            end_color=header_color,
            fill_type='solid'
        )
        header_font = Font(
            bold=True, color='FFFFFF', size=11, name='Calibri'
        )

        # Headers with icons
        headers = [
            '📋 S. No',
            '🏷️ Category',
            '📄 Object Name',
            '🔑 Object UUID',
            '📝 Change Description',
            '💻 Actual SAIL Change',
            '🤖 AI Summary',
            '⚡ Complexity',
            '⏱️ Est. Time',
            '📌 Notes'
        ]

        # Create header row
        ws.row_dimensions[1].height = 35
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            cell.border = thin_border

        # Add changes with alternating colors and category-based highlighting
        changes = data['changes']
        for idx, change in enumerate(changes, 2):
            # Determine row background color
            classification = change['classification']
            if classification == 'CONFLICT':
                row_color = conflict_color
            elif classification == 'NO_CONFLICT':
                row_color = no_conflict_color
            elif classification == 'NEW':
                row_color = new_color
            elif classification == 'DELETED':
                row_color = deleted_color
            else:
                row_color = alt_row_color if idx % 2 == 0 else 'FFFFFF'

            row_fill = PatternFill(
                start_color=row_color, end_color=row_color, fill_type='solid'
            )

            # S. No
            cell = ws.cell(row=idx, column=1)
            cell.value = change['display_order']
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.fill = row_fill
            cell.border = thin_border

            # Category with emoji
            cell = ws.cell(row=idx, column=2)
            category_emoji = self._get_category_emoji(classification)
            cell.value = f"{category_emoji} {classification}"
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.fill = row_fill
            cell.border = thin_border

            # Object Name
            cell = ws.cell(row=idx, column=3)
            cell.value = change['object_name']
            cell.font = Font(size=10)
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # Object UUID
            cell = ws.cell(row=idx, column=4)
            cell.value = change['object_uuid']
            cell.font = Font(size=9, color='666666')
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # Change Description
            change_desc_parts = [f"Type: {change['object_type']}"]
            if change['vendor_change_type']:
                vendor_change = change['vendor_change_type']
                change_desc_parts.append(f"Vendor: {vendor_change}")
            if change['customer_change_type']:
                customer_change = change['customer_change_type']
                change_desc_parts.append(f"Customer: {customer_change}")
            if change['description']:
                change_desc_parts.append(change['description'])
            change_description = ' | '.join(change_desc_parts)

            cell = ws.cell(row=idx, column=5)
            cell.value = change_description
            cell.font = Font(size=10)
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # Actual SAIL Change
            cell = ws.cell(row=idx, column=6)
            if change.get('actual_change'):
                actual_change = change['actual_change']
                # Truncate if too long
                if len(actual_change) > 1000:
                    actual_change = actual_change[:1000] + '\n... [truncated]'
                cell.value = actual_change
                cell.font = Font(size=9, name='Courier New')
            else:
                cell.value = 'N/A'
                cell.font = Font(size=10, italic=True, color='999999')
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # AI Summary
            cell = ws.cell(row=idx, column=7)
            if change.get('ai_summary_status') == 'completed' and change.get('ai_summary'):
                ai_summary = change['ai_summary']
                # Truncate if too long
                if len(ai_summary) > 800:
                    ai_summary = ai_summary[:800] + '\n... [truncated]'
                cell.value = ai_summary
                cell.font = Font(size=10)
            elif change.get('ai_summary_status') == 'processing':
                cell.value = '⏳ Processing...'
                cell.font = Font(size=10, italic=True, color='FFC000')
            elif change.get('ai_summary_status') == 'failed':
                cell.value = '❌ Failed to generate'
                cell.font = Font(size=10, italic=True, color='C00000')
            else:
                cell.value = '⏸️ Pending'
                cell.font = Font(size=10, italic=True, color='999999')
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # Complexity
            complexity = self._calculate_change_complexity(change)
            cell = ws.cell(row=idx, column=8)
            cell.value = complexity
            cell.font = Font(
                bold=True, size=10, color=self._get_complexity_color(
                    complexity
                )
            )
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.fill = row_fill
            cell.border = thin_border

            # Estimated Time
            estimated_time = self._calculate_change_time(complexity)
            cell = ws.cell(row=idx, column=9)
            cell.value = estimated_time
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.fill = row_fill
            cell.border = thin_border

            # Notes
            cell = ws.cell(row=idx, column=10)
            cell.value = change['notes'] or ''
            cell.font = Font(size=10, italic=True)
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=True
            )
            cell.fill = row_fill
            cell.border = thin_border

            # Set row height
            ws.row_dimensions[idx].height = 40

        # Adjust column widths
        column_widths = [10, 18, 30, 38, 45, 50, 45, 14, 14, 35]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    def _get_category_emoji(self, classification):
        """Get emoji for classification category."""
        emoji_map = {
            'CONFLICT': '⚠️',
            'NO_CONFLICT': '✅',
            'NEW': '🆕',
            'DELETED': '🗑️'
        }
        return emoji_map.get(classification, '📌')

    def _get_complexity_color(self, complexity):
        """Get color code for complexity level."""
        color_map = {
            'Low': '70AD47',  # Green
            'Medium': 'FFC000',  # Yellow
            'High': 'FF6B35',  # Orange
            'Critical': 'C00000'  # Red
        }
        return color_map.get(complexity, '000000')
    
    def _calculate_change_complexity(self, change: Dict[str, Any]) -> str:
        """
        Calculate complexity level for a change.
        
        Args:
            change: Change data dictionary
            
        Returns:
            Complexity level: 'Low', 'Medium', 'High', or 'Critical'
        """
        classification = change['classification']
        vendor_change = change['vendor_change_type']
        customer_change = change['customer_change_type']
        
        # CONFLICT is always high or critical
        if classification == 'CONFLICT':
            # If both vendor and customer modified, it's critical
            if vendor_change == 'MODIFIED' and customer_change == 'MODIFIED':
                return 'Critical'
            return 'High'
        
        # NEW objects are typically low complexity
        if classification == 'NEW':
            return 'Low'
        
        # DELETED objects depend on customer changes
        if classification == 'DELETED':
            if customer_change == 'MODIFIED':
                return 'High'
            return 'Medium'
        
        # NO_CONFLICT with modifications
        if classification == 'NO_CONFLICT':
            if vendor_change == 'MODIFIED':
                return 'Medium'
            if vendor_change == 'DEPRECATED':
                return 'Low'
        
        return 'Low'
    
    def _calculate_change_time(self, complexity: str) -> str:
        """
        Calculate estimated time based on complexity.
        
        Args:
            complexity: Complexity level
            
        Returns:
            Estimated time string
        """
        time_mapping = {
            'Low': '15-30 min',
            'Medium': '30-60 min',
            'High': '1-2 hours',
            'Critical': '2-4 hours'
        }
        return time_mapping.get(complexity, '30 min')

    def clear_cache(self, reference_id: Optional[str] = None) -> None:
        """
        Clear cached reports.

        Args:
            reference_id: If provided, clear only reports for this session.
                         If None, clear all cached reports.

        Example:
            >>> service.clear_cache('MS_ABC123')  # Clear specific
            >>> service.clear_cache()  # Clear all cached reports
        """
        if reference_id:
            # Clear specific session reports
            for format in ['xlsx']:
                report_path = self._get_cached_report_path(
                    reference_id, format
                )
                if os.path.exists(report_path):
                    os.remove(report_path)
                    self.logger.info(f"Cleared cached report: {report_path}")
        else:
            # Clear all reports
            if os.path.exists(self.reports_dir):
                for filename in os.listdir(self.reports_dir):
                    file_path = os.path.join(self.reports_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        self.logger.info(
                            f"Cleared cached report: {file_path}"
                        )
